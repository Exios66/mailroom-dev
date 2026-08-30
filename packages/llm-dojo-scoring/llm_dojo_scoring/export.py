"""Excel workbook + codebook export — the canonical experiment-log summary.

Ported from ``scripts/reporting/export_experiment_results.py`` and
``scripts/reporting/export_sweep_results.py`` (llm-entity-extraction) so the
consuming projects can drop their copies and import this one. Column specs
preserve the reference workbook order byte-for-byte (the artifacts in
``Sorter_Experiment_Results.xlsx`` / ``Sorter_Experiment_Codebook.csv`` /
``Sorter_Model_Sweep_Results.xlsx`` were produced by this logic).
"""

from __future__ import annotations

import csv
import json
import os
from datetime import datetime, timezone
from typing import Any, Callable

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .config import PER_SUBTYPE, get_settings

# ---------------------------------------------------------------------------
# Value helpers
# ---------------------------------------------------------------------------


def dotted_get(record: dict, path: str, default: Any = None) -> Any:
    """Dotted-path getter into a record (``scores.sorter.subtype_accuracy``)."""
    cur: Any = record
    for part in path.split("."):
        if isinstance(cur, dict) and part in cur:
            cur = cur[part]
        else:
            return default
    return cur


def display_model(model: str | None) -> str | None:
    """Map a model slug to its display name (fallback: the slug itself)."""
    if not model:
        return None
    return get_settings().model_display.get(model, model)


def display_prompt_version(record: dict) -> str | None:
    """Derive the friendly prompt version (``v12``) from the record."""
    pv = (record.get("prompt_versions") or {}).get("sorter") or record.get("prompt_version")
    if not pv:
        return None
    for prefix in ("sorter_", "contracts_specialist_", "legalbench_task_"):
        if str(pv).startswith(prefix):
            return str(pv)[len(prefix):]
    return pv


def _date(record: dict):
    ts = record.get("timestamp")
    if not ts:
        return None
    try:
        dt = datetime.fromisoformat(str(ts).replace("Z", "+00:00"))
        return dt.astimezone(timezone.utc).replace(hour=0, minute=0, second=0,
                                                   microsecond=0, tzinfo=None)
    except (ValueError, TypeError):
        return None


def _ci(record: dict, path: str, key: str):
    ci = dotted_get(record, path)
    if isinstance(ci, dict):
        return ci.get(key)
    return None


def _tok(record: dict, key: str):
    tok = record.get("tokens") or {}
    if "sorter" in tok and isinstance(tok["sorter"], dict) and key in tok["sorter"]:
        return tok["sorter"].get(key)
    return tok.get(key)


def _tok_str(record: dict, key: str) -> str | None:
    val = dotted_get(record, f"parameters.{key}")
    if val is None:
        return None
    if isinstance(val, bool):
        return "True" if val else "False"
    return str(val)


def _avg_rate(record: dict, path: str):
    d = dotted_get(record, f"scores.{path}")
    if isinstance(d, dict) and d:
        vals = [v for v in d.values() if isinstance(v, (int, float))]
        if not vals:
            return None
        return round(sum(vals) / len(vals), 4)
    return None


# ---------------------------------------------------------------------------
# Format hints
# ---------------------------------------------------------------------------

_PCT = "percent"
_INT = "integer"
_NUM = "number"
_STR = "string"
_DAT = "date"

_FIELDS = ["document_name", "effective_date", "governing_law", "key_obligations",
           "parties", "renewal_terms", "term_length", "termination_clauses"]
_LIST_FIELDS = ["key_obligations", "parties", "termination_clauses"]


def sorter_columns() -> list[dict]:
    """Column spec for the sorter workbook (matches the reference header order)."""
    cols: list[dict] = [
        {"header": "DATE", "get": lambda r: _date(r), "fmt": _DAT,
         "desc": "Date the eval run completed (UTC)", "type": "Date (mm/dd/yyyy)",
         "src": "experiment_log record.timestamp", "example": "08/15/2026"},
        {"header": "Experiment Name", "get": lambda r: r.get("experiment_name"), "fmt": _STR,
         "desc": "Unique run identifier: {model-slug}_{prompt-version}_subtype[_suffix]",
         "type": "String", "src": "experiment_name", "example": "qwen3.7-flash_sorter_v12_subtype_langfuse"},
        {"header": "SAMPLE (n)", "get": lambda r: r.get("n_rows"), "fmt": _INT,
         "desc": "Number of documents evaluated in the run", "type": "Integer",
         "src": "n_rows", "example": "509"},
        {"header": "MODEL", "get": lambda r: display_model(r.get("model")), "fmt": _STR,
         "desc": "LLM that performed classification", "type": "String",
         "src": "model", "example": "Qwen 3.7-Flash"},
        {"header": "Prompt Version", "get": display_prompt_version, "fmt": _STR,
         "desc": "Sorter prompt version tested", "type": "String",
         "src": "prompt_versions.sorter", "example": "v12"},
        {"header": "Temperature", "get": lambda r: dotted_get(r, "parameters.temperature"), "fmt": _NUM,
         "desc": "LLM sampling temperature", "type": "Number",
         "src": "parameters.temperature", "example": "0.1"},
        {"header": "Doc Type Accuracy", "get": lambda r: dotted_get(r, "scores.sorter.exact_match"), "fmt": _PCT,
         "desc": "Accuracy of doc_type classification (contract vs other)", "type": "Percent",
         "src": "scores.sorter.exact_match", "example": "0.9961 (99.61%)"},
        {"header": "Subtype Accuracy", "get": lambda r: dotted_get(r, "scores.sorter.subtype_accuracy"), "fmt": _PCT,
         "desc": "Strict contract-subtype accuracy vs CUAD-folder ground truth", "type": "Percent",
         "src": "scores.sorter.subtype_accuracy", "example": "0.9234 (92.34%)"},
        {"header": "Subtype Accuracy (equiv)", "get": lambda r: dotted_get(r, "scores.sorter.subtype_accuracy_equiv"), "fmt": _PCT,
         "desc": "Family-level subtype accuracy honoring SUBTYPE_EQUIVALENCES (reseller<->distributor, maintenance<->license, development<->license, affiliate<->joint_venture)",
         "type": "Percent", "src": "scores.sorter.subtype_accuracy_equiv", "example": "0.9312 (93.12%)"},
        {"header": "Average Confidence", "get": lambda r: dotted_get(r, "scores.sorter.confidence"), "fmt": _PCT,
         "desc": "Mean model confidence across documents", "type": "Percent",
         "src": "scores.sorter.confidence", "example": "0.9556 (95.56%)"},
        {"header": "Rows Completed", "get": lambda r: r.get("n_ok"), "fmt": _INT,
         "desc": "Documents successfully classified (no parse/API error)", "type": "Integer",
         "src": "n_ok", "example": "509"},
        {"header": "Rows Failed", "get": lambda r: (r.get("n_rows") or 0) - (r.get("n_ok") or 0), "fmt": _INT,
         "desc": "Documents that failed (n_rows - n_ok)", "type": "Integer",
         "src": "derived", "example": "0"},
        {"header": "Subtype Accuracy CI (lo)", "get": lambda r: _ci(r, "scores.sorter.subtype_accuracy_ci", "lo"), "fmt": _PCT,
         "desc": "Percentile-bootstrap 95% CI lower bound for strict subtype accuracy", "type": "Percent",
         "src": "scores.sorter.subtype_accuracy_ci.lo", "example": "0.8978"},
        {"header": "Subtype Accuracy CI (hi)", "get": lambda r: _ci(r, "scores.sorter.subtype_accuracy_ci", "hi"), "fmt": _PCT,
         "desc": "Percentile-bootstrap 95% CI upper bound for strict subtype accuracy", "type": "Percent",
         "src": "scores.sorter.subtype_accuracy_ci.hi", "example": "0.945"},
        {"header": "Subtype Accuracy CI (half)", "get": lambda r: _ci(r, "scores.sorter.subtype_accuracy_ci", "half"), "fmt": _PCT,
         "desc": "Half-width of the subtype accuracy 95% CI", "type": "Percent",
         "src": "scores.sorter.subtype_accuracy_ci.half", "example": "0.0236"},
        {"header": "Doc Type Accuracy CI (lo)", "get": lambda r: _ci(r, "scores.sorter.exact_match_ci", "lo"), "fmt": _PCT,
         "desc": "Percentile-bootstrap 95% CI lower bound for doc_type accuracy", "type": "Percent",
         "src": "scores.sorter.exact_match_ci.lo", "example": "0.9902"},
        {"header": "Doc Type Accuracy CI (hi)", "get": lambda r: _ci(r, "scores.sorter.exact_match_ci", "hi"), "fmt": _PCT,
         "desc": "Percentile-bootstrap 95% CI upper bound for doc_type accuracy", "type": "Percent",
         "src": "scores.sorter.exact_match_ci.hi", "example": "1.0"},
        {"header": "Doc Type Accuracy CI (half)", "get": lambda r: _ci(r, "scores.sorter.exact_match_ci", "half"), "fmt": _PCT,
         "desc": "Half-width of the doc_type accuracy 95% CI", "type": "Percent",
         "src": "scores.sorter.exact_match_ci.half", "example": "0.0049"},
        {"header": "Failures: n Failed", "get": lambda r: dotted_get(r, "scores.sorter.failure_insights.n_failed"), "fmt": _INT,
         "desc": "Total rows misclassified (subtype_ok=False)", "type": "Integer",
         "src": "scores.sorter.failure_insights.n_failed", "example": "39"},
    ]
    for mode, desc in [
        ("equivalent_family", "Failures recovered by subtype equivalence mapping"),
        ("family_confusion", "Misrouted to a semantically distinct family (no equivalence)"),
        ("function_over_form", "doc_type miss (non-contract parsed as contract or vice versa)"),
        ("other_fallback", "Fallback / low-confidence classification failures"),
    ]:
        example = {"equivalent_family": "4", "family_confusion": "29",
                   "function_over_form": "2", "other_fallback": "4"}[mode]
        cols.append({"header": f"Failures: {mode}",
                     "get": lambda r, m=mode: dotted_get(r, f"scores.sorter.failure_insights.mode_counts.{m}"),
                     "fmt": _INT, "desc": desc, "type": "Integer",
                     "src": f"failure_insights.mode_counts.{mode}", "example": example})
    for sub in PER_SUBTYPE:
        cols.append({"header": f"Accuracy: {sub}",
                     "get": lambda r, s=sub: dotted_get(r, f"scores.sorter.per_subtype.{s}.accuracy"),
                     "fmt": _PCT, "desc": f"Strict accuracy for the {sub} subtype class",
                     "type": "Percent", "src": f"scores.sorter.per_subtype.{sub}.accuracy", "example": "0-1"})
    for sub in PER_SUBTYPE:
        cols.append({"header": f"Accuracy (equiv): {sub}",
                     "get": lambda r, s=sub: dotted_get(r, f"scores.sorter.per_subtype.{s}.accuracy_equiv"),
                     "fmt": _PCT, "desc": f"Family-level accuracy for the {sub} subtype class",
                     "type": "Percent", "src": f"scores.sorter.per_subtype.{sub}.accuracy_equiv", "example": "0-1"})
    for sub in PER_SUBTYPE:
        cols.append({"header": f"n Total: {sub}",
                     "get": lambda r, s=sub: dotted_get(r, f"scores.sorter.per_subtype.{s}.total"),
                     "fmt": _INT, "desc": f"Number of documents whose ground-truth subtype is {sub}",
                     "type": "Integer", "src": f"scores.sorter.per_subtype.{sub}.total", "example": "e.g. 34"})
    cols.extend([
        {"header": "Prompt Tokens", "get": lambda r: _tok(r, "prompt_tokens"), "fmt": _INT,
         "desc": "Total input tokens across all docs in the run", "type": "Integer",
         "src": "tokens.sorter.prompt_tokens", "example": "6,712,981"},
        {"header": "Completion Tokens", "get": lambda r: _tok(r, "completion_tokens"), "fmt": _INT,
         "desc": "Total output tokens across all docs in the run", "type": "Integer",
         "src": "tokens.sorter.completion_tokens", "example": "529,692"},
        {"header": "Total Tokens", "get": lambda r: _tok(r, "total_tokens"), "fmt": _INT,
         "desc": "Prompt + completion tokens", "type": "Integer",
         "src": "tokens.sorter.total_tokens", "example": "7,242,673"},
        {"header": "Cost USD", "get": lambda r: _tok(r, "cost_usd"), "fmt": _NUM,
         "desc": "Billed cost (0 when provider not billed through the runner)", "type": "Number",
         "src": "tokens.sorter.cost_usd", "example": "0"},
        {"header": "Cost Total USD", "get": lambda r: _tok(r, "cost_total_usd"), "fmt": _NUM,
         "desc": "Billed total cost bucket", "type": "Number",
         "src": "tokens.sorter.cost_total_usd", "example": "0"},
        {"header": "Cost Estimated USD", "get": lambda r: _tok(r, "cost_estimated_usd"), "fmt": _NUM,
         "desc": "Estimated cost from local cost model (offline)", "type": "Number",
         "src": "tokens.sorter.cost_estimated_usd", "example": "0.270249"},
        {"header": "Rows with Usage", "get": lambda r: _tok(r, "rows_with_usage"), "fmt": _INT,
         "desc": "Docs with token/usage metadata (manifest-replayed rows carry none)", "type": "Integer",
         "src": "tokens.sorter.rows_with_usage", "example": "509"},
        {"header": "Reasoning Effort", "get": lambda r: dotted_get(r, "parameters.reasoning_effort"), "fmt": _STR,
         "desc": "reasoning_effort parameter for the sorter (default medium)", "type": "String",
         "src": "parameters.reasoning_effort", "example": "medium"},
        {"header": "Max Tokens", "get": lambda r: dotted_get(r, "parameters.max_tokens"), "fmt": _INT,
         "desc": "LLM max output tokens", "type": "Integer", "src": "parameters.max_tokens", "example": "4096"},
        {"header": "Max Input Chars", "get": lambda r: dotted_get(r, "parameters.max_input_chars"), "fmt": _INT,
         "desc": "Max input characters before head+tail truncation", "type": "Integer",
         "src": "parameters.max_input_chars", "example": "100000"},
        {"header": "Max Concurrency", "get": lambda r: dotted_get(r, "parameters.max_concurrency"), "fmt": _INT,
         "desc": "Parallel document concurrency", "type": "Integer",
         "src": "parameters.max_concurrency", "example": "8"},
        {"header": "Dataset", "get": lambda r: dotted_get(r, "data_source.project"), "fmt": _STR,
         "desc": "Braintrust dataset project used", "type": "String",
         "src": "data_source.project", "example": "llm-mailroom/mailroom-cuad-contracts-full"},
        {"header": "n Samples", "get": lambda r: dotted_get(r, "data_source.n_samples"), "fmt": _INT,
         "desc": "Documents in the dataset used", "type": "Integer",
         "src": "data_source.n_samples", "example": "509"},
        {"header": "Sample Requested", "get": lambda r: dotted_get(r, "data_source.sample_requested"), "fmt": _INT,
         "desc": "Requested sample size (0 = full dataset)", "type": "Integer",
         "src": "data_source.sample_requested", "example": "0"},
        {"header": "Stratified", "get": lambda r: dotted_get(r, "data_source.stratified"), "fmt": _STR,
         "desc": "Stratified sampling per subtype (200 = 8/subtype x 25)", "type": "Integer/None",
         "src": "data_source.stratified", "example": "0"},
        {"header": "Seed", "get": lambda r: dotted_get(r, "data_source.seed"), "fmt": _INT,
         "desc": "Sampling seed (same-seed runs are directly comparable)", "type": "Integer",
         "src": "data_source.seed", "example": "42"},
    ])
    return cols


def extraction_columns() -> list[dict]:
    """Column spec for the extraction workbook (matches the reference header order)."""
    cols: list[dict] = [
        {"header": "DATE", "get": lambda r: _date(r), "fmt": _DAT,
         "desc": "Date the eval run completed (UTC)", "type": "Date (mm/dd/yyyy)",
         "src": "experiment_log record.timestamp", "example": "08/15/2026"},
        {"header": "Experiment Name", "get": lambda r: r.get("experiment_name"), "fmt": _STR,
         "desc": "Unique run identifier: {model-slug}_{prompt-version}_extraction[_suffix]",
         "type": "String", "src": "experiment_name",
         "example": "qwen3.7-flash_contracts_specialist_v32_extraction_langfuse_510_full"},
        {"header": "SAMPLE (n)", "get": lambda r: r.get("n_rows"), "fmt": _INT,
         "desc": "Number of documents evaluated in the run", "type": "Integer",
         "src": "n_rows", "example": "509"},
        {"header": "MODEL", "get": lambda r: display_model(r.get("model")), "fmt": _STR,
         "desc": "LLM that performed extraction", "type": "String",
         "src": "model", "example": "Qwen 3.7-Flash"},
        {"header": "Prompt Version", "get": display_prompt_version, "fmt": _STR,
         "desc": "Contracts-specialist prompt version tested", "type": "String",
         "src": "prompt_version", "example": "v32"},
        {"header": "Temperature", "get": lambda r: dotted_get(r, "parameters.temperature"), "fmt": _NUM,
         "desc": "LLM sampling temperature", "type": "Number",
         "src": "parameters.temperature", "example": "0.1"},
        {"header": "Overall Extraction", "get": lambda r: dotted_get(r, "scores.overall_extraction_score"), "fmt": _PCT,
         "desc": "Composite extraction score (field-type-aware content accuracy)", "type": "Percent",
         "src": "scores.overall_extraction_score", "example": "0.887"},
        {"header": "Field Presence", "get": lambda r: dotted_get(r, "scores.field_presence"), "fmt": _PCT,
         "desc": "Fraction of expected fields present in the extraction", "type": "Percent",
         "src": "scores.field_presence", "example": "0.9728"},
        {"header": "Schema Valid?", "get": lambda r: dotted_get(r, "scores.schema_valid"), "fmt": _PCT,
         "desc": "Fraction of rows whose extraction parsed to a valid schema", "type": "Percent",
         "src": "scores.schema_valid", "example": "1.0"},
    ]
    for f in ["doc_name", "effective_date", "governing_law", "key_obligations", "Parties",
              "Renewal Terms", "Term_Length", "Termination_Clauses"]:
        key = {"doc_name": "document_name", "Parties": "parties", "Renewal Terms": "renewal_terms",
               "Term_Length": "term_length", "Termination_Clauses": "termination_clauses"}.get(f, f)
        cols.append({"header": f, "get": lambda r, k=key: dotted_get(r, f"scores.per_field.{k}"), "fmt": _PCT,
                     "desc": f"Per-field score for {key}", "type": "Percent",
                     "src": f"scores.per_field.{key}", "example": "0-1"})
    cols.extend([
        {"header": "Overall Extraction CI (lo)", "get": lambda r: _ci(r, "scores.overall_extraction_score_ci", "lo"), "fmt": _PCT,
         "desc": "Percentile-bootstrap 95% CI lower bound for the composite score", "type": "Percent",
         "src": "scores.overall_extraction_score_ci.lo", "example": "0.8689"},
        {"header": "Overall Extraction CI (hi)", "get": lambda r: _ci(r, "scores.overall_extraction_score_ci", "hi"), "fmt": _PCT,
         "desc": "Percentile-bootstrap 95% CI upper bound for the composite score", "type": "Percent",
         "src": "scores.overall_extraction_score_ci.hi", "example": "0.8913"},
        {"header": "Overall Extraction CI (half)", "get": lambda r: _ci(r, "scores.overall_extraction_score_ci", "half"), "fmt": _PCT,
         "desc": "Half-width of the composite score 95% CI", "type": "Percent",
         "src": "scores.overall_extraction_score_ci.half", "example": "0.0112"},
        {"header": "Overall Verified Precision", "get": lambda r: dotted_get(r, "scores.overall_verified_precision"), "fmt": _PCT,
         "desc": "Precision over items verified in the document or matched to GT", "type": "Percent",
         "src": "scores.overall_verified_precision", "example": "0.9819"},
        {"header": "Hallucination Rate (avg)", "get": lambda r: _avg_rate(r, "hallucination_rate"), "fmt": _PCT,
         "desc": "Mean hallucination rate across fields (predicted item grounded in neither GT nor doc)",
         "type": "Percent", "src": "scores.hallucination_rate (mean)", "example": "0.0147"},
    ])
    for f in _FIELDS:
        cols.append({"header": f"H Rate: {f}",
                     "get": lambda r, k=f: dotted_get(r, f"scores.hallucination_rate.{k}"), "fmt": _PCT,
                     "desc": f"Hallucination rate for {f}", "type": "Percent",
                     "src": f"scores.hallucination_rate.{f}", "example": "0-1"})
    for f in _FIELDS:
        cols.append({"header": f"Verified Precision: {f}",
                     "get": lambda r, k=f: dotted_get(r, f"scores.verified_precision.{k}"), "fmt": _PCT,
                     "desc": f"Verified precision for {f}", "type": "Percent",
                     "src": f"scores.verified_precision.{f}", "example": "0-1"})
    for f in _LIST_FIELDS:
        cols.append({"header": f"Entity List F1: {f}",
                     "get": lambda r, k=f: dotted_get(r, f"scores.entity_list_f1.{k}"), "fmt": _PCT,
                     "desc": f"Entity-list F1 (bipartite match, GT coverage) for {f}", "type": "Percent",
                     "src": f"scores.entity_list_f1.{f}", "example": "0-1"})
    diag_map = [
        ("Diag: Field Exact Rate", "field_exact_rate", _PCT, "Fraction of scored fields exact"),
        ("Diag: Field Partial Rate", "field_partial_rate", _PCT, "Fraction of scored fields partial"),
        ("Diag: Field Miss Rate", "field_miss_rate", _PCT, "Fraction of scored fields missed"),
        ("Diag: n Fields Scored", "n_fields_scored", _INT, "Number of field scores in the run"),
        ("Diag: List Precision", "list_precision", _PCT, "Raw entity-list precision (macro)"),
        ("Diag: List Recall", "list_recall", _PCT, "Raw entity-list recall (macro)"),
        ("Diag: List F1", "list_f1", _PCT, "Raw entity-list F1 (macro)"),
        ("Diag: List Micro Precision", "list_micro_precision", _PCT, "Micro entity-list precision"),
        ("Diag: List Micro Recall", "list_micro_recall", _PCT, "Micro entity-list recall"),
        ("Diag: List Micro F1", "list_micro_f1", _PCT, "Micro entity-list F1"),
        ("Diag: List Micro Matched", "list_micro_matched", _INT, "Micro matched items"),
        ("Diag: List Micro n Expected", "list_micro_n_expected", _INT, "Micro expected items"),
        ("Diag: List Micro n Predicted", "list_micro_n_predicted", _INT, "Micro predicted items"),
    ]
    for header, key, fmt, desc in diag_map:
        cols.append({"header": header, "get": lambda r, k=key: dotted_get(r, f"scores.diagnostics.{k}"), "fmt": fmt,
                     "desc": desc, "type": "Integer" if fmt == _INT else "Number",
                     "src": f"scores.diagnostics.{key}", "example": "0-1"})
    for f in _LIST_FIELDS:
        cols.append({"header": f"Diag: List Precision: {f}",
                     "get": lambda r, k=f: dotted_get(r, f"scores.diagnostics.entity_list_precision.{k}"), "fmt": _PCT,
                     "desc": f"Raw entity-list precision for {f}", "type": "Percent",
                     "src": f"scores.diagnostics.entity_list_precision.{f}", "example": "0-1"})
    for f in _LIST_FIELDS:
        cols.append({"header": f"Diag: List Recall: {f}",
                     "get": lambda r, k=f: dotted_get(r, f"scores.diagnostics.entity_list_recall.{k}"), "fmt": _PCT,
                     "desc": f"Raw entity-list recall for {f}", "type": "Percent",
                     "src": f"scores.diagnostics.entity_list_recall.{f}", "example": "0-1"})
    for f in _LIST_FIELDS:
        cols.append({"header": f"Diag: List Raw F1: {f}",
                     "get": lambda r, k=f: dotted_get(r, f"scores.diagnostics.entity_list_raw_f1.{k}"), "fmt": _PCT,
                     "desc": f"Raw entity-list F1 for {f}", "type": "Percent",
                     "src": f"scores.diagnostics.entity_list_raw_f1.{f}", "example": "0-1"})
    for header, key, fmt, desc in [
        ("Diag: Date MAE (days)", "date_mae_days", _NUM, "Mean absolute error over date pairs (days)"),
        ("Diag: Date Median AE (days)", "date_median_ae_days", _NUM, "Median absolute error over date pairs (days)"),
        ("Diag: Date R2", "date_r2", _NUM, "R2 for date predictions"),
        ("Diag: Date n pairs", "date_n_pairs", _INT, "Number of parseable date pairs"),
    ]:
        cols.append({"header": header, "get": lambda r, k=key: dotted_get(r, f"scores.diagnostics.{k}"), "fmt": fmt,
                     "desc": desc, "type": "Integer" if fmt == _INT else "Number",
                     "src": f"scores.diagnostics.{key}", "example": "0-1"})
    for f in ["effective_date", "term_length"]:
        cols.append({"header": f"Diag: Date MAE: {f}",
                     "get": lambda r, k=f: dotted_get(r, f"scores.diagnostics.date_mae_per_field.{k}"), "fmt": _NUM,
                     "desc": f"Date MAE for {f} (days)", "type": "Number",
                     "src": f"scores.diagnostics.date_mae_per_field.{f}", "example": "0-365"})
    for f in ["effective_date", "term_length"]:
        cols.append({"header": f"Diag: Date R2: {f}",
                     "get": lambda r, k=f: dotted_get(r, f"scores.diagnostics.date_r2_per_field.{k}"), "fmt": _NUM,
                     "desc": f"Date R2 for {f}", "type": "Number",
                     "src": f"scores.diagnostics.date_r2_per_field.{f}", "example": "0-1"})
    for header, key, fmt, desc in [
        ("Diag: Duration MAE (days)", "duration_mae_days", _NUM, "Mean absolute error over duration pairs (days)"),
        ("Diag: Duration Median AE (days)", "duration_median_ae_days", _NUM, "Median absolute error over duration pairs (days)"),
        ("Diag: Duration R2", "duration_r2", _NUM, "R2 for duration predictions"),
        ("Diag: Duration n pairs", "duration_n_pairs", _INT, "Number of parseable duration pairs"),
    ]:
        cols.append({"header": header, "get": lambda r, k=key: dotted_get(r, f"scores.diagnostics.{k}"), "fmt": fmt,
                     "desc": desc, "type": "Integer" if fmt == _INT else "Number",
                     "src": f"scores.diagnostics.{key}", "example": "0-1"})
    for f in ["renewal_terms", "term_length"]:
        cols.append({"header": f"Diag: Duration MAE: {f}",
                     "get": lambda r, k=f: dotted_get(r, f"scores.diagnostics.duration_mae_per_field.{k}"), "fmt": _NUM,
                     "desc": f"Duration MAE for {f} (days)", "type": "Number",
                     "src": f"scores.diagnostics.duration_mae_per_field.{f}", "example": "0-365"})
    for f in ["renewal_terms", "term_length"]:
        cols.append({"header": f"Diag: Duration R2: {f}",
                     "get": lambda r, k=f: dotted_get(r, f"scores.diagnostics.duration_r2_per_field.{k}"), "fmt": _NUM,
                     "desc": f"Duration R2 for {f}", "type": "Number",
                     "src": f"scores.diagnostics.duration_r2_per_field.{f}", "example": "0-1"})
    for header, key, fmt, desc in [
        ("Diag: Money n pairs", "money_n_pairs", _INT, "Number of parseable money pairs"),
        ("Diag: Span Count MAE", "span_count_mae", _NUM, "Mean absolute span-count error"),
        ("Diag: Span Count n docs", "span_count_n_docs", _INT, "Documents with span-count diagnostics"),
        ("Diag: Span Count Signed Mean", "span_count_signed_mean", _NUM, "Signed mean span-count drift (over vs under extraction)"),
    ]:
        cols.append({"header": header, "get": lambda r, k=key: dotted_get(r, f"scores.diagnostics.{k}"), "fmt": fmt,
                     "desc": desc, "type": "Integer" if fmt == _INT else "Number",
                     "src": f"scores.diagnostics.{key}", "example": "0-1"})
    for f in _LIST_FIELDS:
        cols.append({"header": f"Diag: Span MAE: {f}",
                     "get": lambda r, k=f: dotted_get(r, f"scores.diagnostics.span_count_mae_per_field.{k}"), "fmt": _NUM,
                     "desc": f"Span-count MAE for {f}", "type": "Number",
                     "src": f"scores.diagnostics.span_count_mae_per_field.{f}", "example": "0-10"})
    for f in _LIST_FIELDS:
        cols.append({"header": f"Diag: Span Signed Mean: {f}",
                     "get": lambda r, k=f: dotted_get(r, f"scores.diagnostics.span_count_signed_mean_per_field.{k}"), "fmt": _NUM,
                     "desc": f"Span-count signed mean for {f} (over-extraction positive)", "type": "Number",
                     "src": f"scores.diagnostics.span_count_signed_mean_per_field.{f}", "example": "0-10"})
    for f in _FIELDS + ["contract_value"]:
        cols.append({"header": f"Diag: Field Presence: {f}",
                     "get": lambda r, k=f: dotted_get(r, f"scores.diagnostics.field_presence_per_field.{k}"), "fmt": _PCT,
                     "desc": f"Field-presence rate for {f}", "type": "Percent",
                     "src": f"scores.diagnostics.field_presence_per_field.{f}", "example": "0-1"})
    for f in _FIELDS:
        for part, pkey in [("exact_rate", "exact_rate"), ("partial_rate", "partial_rate"), ("miss_rate", "miss_rate")]:
            cols.append({"header": f"Err {part} : {f}",
                         "get": lambda r, k=f, pk=pkey: dotted_get(r, f"scores.diagnostics.error_decomposition.{k}.{pk}"),
                         "fmt": _PCT, "desc": f"Error decomposition {part} for {f}", "type": "Percent",
                         "src": f"scores.diagnostics.error_decomposition.{f}.{pkey}", "example": "0-1"})
    cols.append({"header": "Category Presence", "get": lambda r: dotted_get(r, "scores.category_presence"), "fmt": _PCT,
                 "desc": "CUAD category presence rate (expected categories found)", "type": "Percent",
                 "src": "scores.category_presence", "example": "0.8555"})
    cols.extend([
        {"header": "Prompt Tokens", "get": lambda r: _tok(r, "prompt_tokens"), "fmt": _INT,
         "desc": "Total input tokens across all docs in the run", "type": "Integer",
         "src": "tokens.prompt_tokens", "example": "6,712,981"},
        {"header": "Completion Tokens", "get": lambda r: _tok(r, "completion_tokens"), "fmt": _INT,
         "desc": "Total output tokens across all docs in the run", "type": "Integer",
         "src": "tokens.completion_tokens", "example": "529,692"},
        {"header": "Total Tokens", "get": lambda r: _tok(r, "total_tokens"), "fmt": _INT,
         "desc": "Prompt + completion tokens", "type": "Integer",
         "src": "tokens.total_tokens", "example": "7,242,673"},
        {"header": "Cost USD", "get": lambda r: _tok(r, "cost_usd"), "fmt": _NUM,
         "desc": "Billed cost (0 when provider not billed through the runner)", "type": "Number",
         "src": "tokens.cost_usd", "example": "0"},
        {"header": "Cost Total USD", "get": lambda r: _tok(r, "cost_total_usd"), "fmt": _STR,
         "desc": "Billed total cost bucket (string in the reference extraction workbook)",
         "type": "String", "src": "tokens.cost_total_usd", "example": "0"},
        {"header": "Cost Estimated USD", "get": lambda r: _tok(r, "cost_estimated_usd"), "fmt": _NUM,
         "desc": "Estimated cost from local cost model (offline)", "type": "Number",
         "src": "tokens.cost_estimated_usd", "example": "0.270249"},
        {"header": "Rows with Usage", "get": lambda r: _tok(r, "rows_with_usage"), "fmt": _INT,
         "desc": "Docs with token/usage metadata (manifest-replayed rows carry none)", "type": "Integer",
         "src": "tokens.rows_with_usage", "example": "509"},
        {"header": "Reasoning Effort", "get": lambda r: dotted_get(r, "parameters.reasoning_effort"), "fmt": _STR,
         "desc": "reasoning_effort parameter for the extractor (default none)", "type": "String",
         "src": "parameters.reasoning_effort", "example": "none"},
        {"header": "Max Tokens", "get": lambda r: _tok_str(r, "max_tokens"), "fmt": _STR,
         "desc": "LLM max output tokens (string in the reference extraction workbook)",
         "type": "String", "src": "parameters.max_tokens", "example": "16384"},
        {"header": "Max Input Chars", "get": lambda r: dotted_get(r, "parameters.max_input_chars"), "fmt": _INT,
         "desc": "Max input characters before head+tail truncation", "type": "Integer",
         "src": "parameters.max_input_chars", "example": "150000"},
        {"header": "Max Concurrency", "get": lambda r: dotted_get(r, "parameters.max_concurrency"), "fmt": _INT,
         "desc": "Parallel document concurrency", "type": "Integer",
         "src": "parameters.max_concurrency", "example": "8"},
        {"header": "Chunked", "get": lambda r: (int(bool(dotted_get(r, "parameters.chunked")))
                                                if dotted_get(r, "parameters.chunked") is not None else None),
         "fmt": _INT, "desc": "Whether the run used chunked extraction (1 = yes, 0 = no)",
         "type": "Integer", "src": "parameters.chunked", "example": "1"},
        {"header": "Chunk Chars", "get": lambda r: _tok_str(r, "chunk_chars"), "fmt": _STR,
         "desc": "Chunk window size in characters (string in the reference extraction workbook)",
         "type": "String", "src": "parameters.chunk_chars", "example": "90000"},
        {"header": "Chunk Overlap", "get": lambda r: dotted_get(r, "parameters.chunk_overlap"), "fmt": _INT,
         "desc": "Chunk overlap in characters (chunked runs)", "type": "Integer",
         "src": "parameters.chunk_overlap", "example": "8000"},
        {"header": "Judge", "get": lambda r: _tok_str(r, "judge"), "fmt": _STR,
         "desc": "Whether the judge LLM pass was enabled (string in the reference extraction workbook)",
         "type": "Boolean/String", "src": "parameters.judge", "example": "False"},
        {"header": "Dataset", "get": lambda r: dotted_get(r, "data_source.project"), "fmt": _STR,
         "desc": "Braintrust dataset project used", "type": "String",
         "src": "data_source.project", "example": "llm-mailroom/mailroom-cuad-contracts-full"},
        {"header": "n Samples", "get": lambda r: dotted_get(r, "data_source.n_samples"), "fmt": _INT,
         "desc": "Documents in the dataset used", "type": "Integer",
         "src": "data_source.n_samples", "example": "509"},
        {"header": "Seed", "get": lambda r: dotted_get(r, "data_source.seed"), "fmt": _INT,
         "desc": "Sampling seed (same-seed runs are directly comparable)", "type": "Integer",
         "src": "data_source.seed", "example": "42"},
    ])
    return cols


# ---------------------------------------------------------------------------
# Workbook / codebook writers (Google-Sheets-friendly)
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FORMATS = {_PCT: "0.00%", _DAT: "mm/dd/yyyy", _INT: "#,##0", _NUM: "0.000000", _STR: "General"}

_COL_WIDTHS = {"DATE": 12, "Experiment Name": 46, "SAMPLE (n)": 10, "MODEL": 20,
               "Prompt Version": 13, "Temperature": 11, "Dataset": 42}


def _width(header: str) -> float:
    if header in _COL_WIDTHS:
        return _COL_WIDTHS[header]
    if header.startswith("Accuracy") or header.startswith("H Rate") or header.startswith("Diag"):
        return 15 if len(header) < 22 else 22
    if header.startswith("n Total"):
        return 15
    return 12


def write_workbook(path: str, sheet_title: str, columns: list[dict], records: list[dict],
                   codebook_sheet: bool = False) -> None:
    """Write the Eval Results workbook with Google-Sheets-friendly styling."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    headers = [c["header"] for c in columns]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER

    for rec in records:
        row = []
        for col in columns:
            row.append(col["get"](rec))
        ws.append(row)

    for cidx, col in enumerate(columns, start=1):
        letter = get_column_letter(cidx)
        ws.column_dimensions[letter].width = _width(col["header"])
        for ridx in range(2, len(records) + 2):
            cell = ws.cell(row=ridx, column=cidx)
            if col["fmt"] == _PCT and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00%"
            elif col["fmt"] == _DAT and cell.value is not None:
                cell.number_format = "mm/dd/yyyy"
            elif col["fmt"] == _INT and isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"
            elif col["fmt"] == _NUM and isinstance(cell.value, (int, float)):
                cell.number_format = "0.000000"
            cell.border = _BORDER

    ws.freeze_panes = "F2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(records) + 1}"
    if codebook_sheet:
        write_compact_codebook_sheet(wb, columns)
    wb.save(path)


def write_codebook(path: str, columns: list[dict]) -> None:
    """Write the full variable dictionary as a Google-Sheets-friendly CSV."""
    with open(path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["Variable", "Description", "Type", "Source", "Example / Values"])
        for col in columns:
            writer.writerow([col["header"], col["desc"], col["type"], col["src"], col["example"]])


def write_compact_codebook_sheet(wb: openpyxl.Workbook, columns: list[dict]) -> None:
    """Add the compact Codebook sheet (one row per variable) to the workbook."""
    ws = wb.create_sheet("Codebook")
    ws.append(["Variable", "Description", "Type", "Source", "Example / Values"])
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    for col in columns:
        ws.append([col["header"], col["desc"], col["type"], col["src"], col["example"]])
    for cidx, width in [(1, 30), (2, 90), (3, 16), (4, 42), (5, 24)]:
        ws.column_dimensions[get_column_letter(cidx)].width = width
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Task filters + sweep builder
# ---------------------------------------------------------------------------

def sorter_records(records: list[dict]) -> list[dict]:
    return [r for r in records if r.get("task") == "subtype_classification"]


def extraction_records(records: list[dict]) -> list[dict]:
    return [r for r in records if r.get("task") == "contract_entity_extraction"]


def load_records(log_path: str) -> list[dict]:
    """Load experiment-log records (chronological order preserved)."""
    records = []
    with open(log_path, encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if not line:
                continue
            records.append(json.loads(line))
    return records


def champion_prompt_version(records: list[dict], task: str = "subtype_classification",
                            metric_path: str = "scores.sorter.subtype_accuracy") -> str | None:
    """Prompt version of the best-scoring record (ties: larger n) — the
    champion filter used by the sweep workbook."""
    candidates = [r for r in records if r.get("task") == task]
    if not candidates:
        return None

    def score(r: dict) -> float:
        val = dotted_get(r, metric_path)
        try:
            return float(val or 0.0)
        except (TypeError, ValueError):
            return 0.0

    return display_prompt_version(max(candidates, key=lambda r: (score(r), r.get("n_rows") or 0)))


def sweep_records(records: list[dict], task: str = "subtype_classification",
                  prompt_version: str | None = None,
                  notes: dict[str, str] | None = None) -> list[dict]:
    """Records for the model-sweep workbook: champion-prompt runs (all models)
    with an optional per-experiment Notes column attached."""
    if prompt_version is None:
        prompt_version = champion_prompt_version(records, task)
    out: list[dict] = []
    for record in records:
        if record.get("task") != task:
            continue
        if display_prompt_version(record) != prompt_version:
            continue
        record = dict(record)
        note = (notes or {}).get(record.get("experiment_name"))
        if note:
            record["_note"] = note
        out.append(record)
    return out


def build_sweep_workbook(records: list[dict], outdir: str = ".",
                         prompt_version: str | None = None,
                         notes: dict[str, str] | None = None) -> tuple[str, int]:
    """Write the sweep workbook + codebook; returns (path, n_records)."""
    recs = sweep_records(records, prompt_version=prompt_version, notes=notes)
    columns = sorter_columns()
    if any(r.get("_note") for r in recs):
        columns = columns + [{
            "header": "Notes", "get": lambda r: r.get("_note"), "fmt": _STR,
            "desc": "Curated note for the sweep row", "type": "String",
            "src": "_note", "example": "rerun with usage",
        }]
    path = os.path.join(outdir, "Sorter_Model_Sweep_Results.xlsx")
    write_workbook(path, "Eval Results", columns, recs, codebook_sheet=True)
    write_codebook(os.path.join(outdir, "Sorter_Experiment_Codebook.csv"), columns)
    return path, len(recs)
